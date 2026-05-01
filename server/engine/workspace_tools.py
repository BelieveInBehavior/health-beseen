"""
Workspace tools — actone-ai 风格：`list_files` / `read_file` / `write_file` / `delete_file` /
`read_document`（Office 可选依赖）+ `bash`。

路径限制在 HEALTH_BESEEN_WORKSPACE。Skill 索引见 `skills_index.load_skills_index_markdown()`（注入 system）。
"""

from __future__ import annotations

import fnmatch
import os
import re
import subprocess
from pathlib import Path

from server.config import settings
from server.engine.skills_index import scan_skills_catalog

_DEFAULT_LINE_LIMIT = 500


def _rel_to_root(fp: Path) -> str:
    root = _workspace_root()
    try:
        return str(fp.relative_to(root))
    except ValueError:
        return str(fp)


def _workspace_root() -> Path:
    return Path(settings.WORKSPACE_ROOT).resolve()


def resolve_workspace_path(user_path: str) -> Path:
    """Resolve *user_path* under workspace; raise ValueError if escape."""
    root = _workspace_root()
    p = Path(user_path.strip()).expanduser()
    if not p.is_absolute():
        p = (root / p).resolve()
    else:
        p = p.resolve()
    try:
        p.relative_to(root)
    except ValueError as e:
        raise ValueError(f"路径必须在项目工作区内: {user_path}") from e
    return p


def resolve_cwd_relative(rel: str | None) -> Path:
    """Resolve optional cwd (relative to workspace) to absolute path under workspace."""
    root = _workspace_root()
    if not rel or not str(rel).strip():
        return root
    sub = Path(rel.strip())
    if sub.is_absolute():
        p = sub.resolve()
    else:
        p = (root / sub).resolve()
    try:
        p.relative_to(root)
    except ValueError as e:
        raise ValueError(f"cwd 必须在项目工作区内: {rel}") from e
    if not p.is_dir():
        raise ValueError(f"cwd 不是目录: {p}")
    return p


def _blocked_device_path(path: Path) -> bool:
    s = str(path)
    if s in {
        "/dev/zero", "/dev/random", "/dev/urandom", "/dev/full",
        "/dev/stdin", "/dev/stdout", "/dev/stderr", "/dev/tty", "/dev/console",
    }:
        return True
    parts = path.parts
    if len(parts) >= 2 and parts[0] == "/" and parts[1] == "dev":
        return True
    if "/proc/" in s or s.startswith("/sys/"):
        return True
    return False


def read_file_tool(
    path: str,
    offset: int = 0,
    limit: int | None = None,
) -> dict:
    """
    Read a text file; *offset* is 0-based line index (router schema).
    Returns dict for chat SSE formatting.
    """
    lim = int(limit) if limit is not None else _DEFAULT_LINE_LIMIT
    lim = max(1, min(lim, 2000))
    try:
        off = max(0, int(offset))
    except (TypeError, ValueError):
        off = 0

    out: dict = {"ok": False, "path": path or ""}

    if not path or not str(path).strip():
        out["error"] = "path 为空"
        return out

    try:
        fp = resolve_workspace_path(path)
    except ValueError as e:
        out["error"] = str(e)
        return out

    if _blocked_device_path(fp):
        out["error"] = "拒绝读取该路径"
        return out

    if not fp.is_file():
        out["error"] = f"不是文件或不存在: {fp}"
        return out

    try:
        raw = fp.read_bytes()
    except OSError as e:
        out["error"] = str(e)
        return out

    max_bytes = max(4096, settings.READ_FILE_MAX_BYTES)
    truncated_bytes = len(raw) > max_bytes
    if truncated_bytes:
        raw = raw[:max_bytes]

    if b"\x00" in raw[:8192]:
        out["error"] = "疑似二进制文件"
        return out

    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")

    lines = text.splitlines()
    total = len(lines)
    end = min(total, off + lim)
    if off >= total:
        chunk: list[str] = []
    else:
        chunk = lines[off:end]

    max_line_len = 2000
    numbered: list[str] = []
    for i, line in enumerate(chunk, start=off + 1):
        short = line if len(line) <= max_line_len else line[:max_line_len] + "…"
        numbered.append(f"{i:6}|{short}")

    content = "\n".join(numbered)
    out.update({
        "ok": True,
        "path": _rel_to_root(fp),
        "offset": off,
        "lines": len(chunk),
        "total_lines": total,
        "content": content,
        "truncated_bytes": truncated_bytes,
    })
    return out


_BASH_BLOCK_PATTERNS = [
    re.compile(r"^\s*sudo\b", re.I),
    re.compile(r";\s*sudo\b", re.I),
    re.compile(r"\b(?:curl|wget)\s+[^\n]*\|\s*(?:bash|sh)\b", re.I),
    re.compile(r":\(\)\s*\{\s*:\|:&\s*\}\s*;:", re.I),
    re.compile(r"\bchmod\s+[-+]?[rwxXst]+\s+/\s*$", re.I),
    re.compile(r"\brm\s+[^\n]*\s+/(?:\s|$)", re.I),
]


def bash_tool(command: str, cwd: str | None = None) -> dict:
    """Run bash -lc under workspace (or subdirectory)."""
    if not settings.ENABLE_BASH_TOOL:
        return {"ok": False, "error": "bash 工具已在服务器禁用（ENABLE_BASH_TOOL=0）", "cwd": ".", "output": ""}

    cmd = (command or "").strip()
    root = _workspace_root()

    if not cmd:
        return {"ok": False, "error": "command 为空", "cwd": ".", "output": ""}

    for pat in _BASH_BLOCK_PATTERNS:
        if pat.search(cmd):
            return {"ok": False, "error": "出于安全考虑，该命令已被拒绝。", "cwd": ".", "output": ""}

    try:
        run_cwd = resolve_cwd_relative(cwd)
    except ValueError as e:
        return {"ok": False, "error": str(e), "cwd": ".", "output": ""}

    rel_cwd = str(run_cwd.relative_to(root)) if run_cwd != root else "."

    timeout = float(settings.WORKSPACE_BASH_TIMEOUT_SEC)
    max_chars = settings.WORKSPACE_BASH_MAX_OUTPUT_CHARS

    try:
        proc = subprocess.run(
            ["/bin/bash", "-lc", cmd],
            cwd=run_cwd,
            capture_output=True,
            text=True,
            timeout=timeout,
            env={**os.environ, "LC_ALL": "C.UTF-8"},
        )
    except subprocess.TimeoutExpired:
        return {
            "ok": False,
            "error": f"超时（>{timeout}s）",
            "cwd": rel_cwd,
            "output": "",
        }
    except OSError as e:
        return {"ok": False, "error": str(e), "cwd": rel_cwd, "output": ""}

    combined = proc.stdout or ""
    if proc.stderr:
        combined += ("\n" if combined and not combined.endswith("\n") else "") + proc.stderr
    if len(combined) > max_chars:
        combined = combined[:max_chars] + f"\n…[截断至 {max_chars} 字符]"

    return {
        "ok": proc.returncode == 0,
        "exit_code": proc.returncode,
        "cwd": rel_cwd,
        "output": combined,
    }


def _skills_dirs() -> list[Path]:
    root = _workspace_root()
    sd = Path(settings.SKILLS_DIR)
    primary = sd if sd.is_absolute() else (root / sd)
    candidates = [primary, root / ".cursor" / "skills"]
    return [p for p in candidates if p.is_dir()]


def list_skills_tool() -> dict:
    """Skill 清单（与 `skills_index.scan_skills_catalog` 同源，一层目录 + 平铺 .md）。"""
    rows = scan_skills_catalog()
    skills = [{"path": r["path"], "name": r["name"], "description": r.get("description", "")} for r in rows]
    return {"ok": True, "skills": skills}


def _grep_skip_dir_names() -> set[str]:
    return {x.strip() for x in settings.WORKSPACE_GREP_SKIP_DIRS.split(",") if x.strip()}


def glob_files_tool(pattern: str, root: str = "") -> dict:
    """
    在工作区内按 glob 列出路径（Python pathlib，不启 shell）。
    pattern 如 `**/*.py`、`skills/**/*.md`；root 为相对工作区的子目录，默认根。
    """
    pat = (pattern or "").strip()
    if not pat or ".." in pat or pat.startswith(("/", "\\")):
        return {"ok": False, "error": "pattern 非法或为空", "paths": []}

    root_ws = _workspace_root()
    rel_root = (root or "").strip().replace("\\", "/").strip("/")
    try:
        base = (root_ws / rel_root).resolve() if rel_root else root_ws
        base.relative_to(root_ws)
    except ValueError:
        return {"ok": False, "error": "root 必须在工作区内", "paths": []}

    if not base.exists():
        return {"ok": False, "error": f"root 不存在: {rel_root or '.'}", "paths": []}

    try:
        matches = sorted({str(p.relative_to(root_ws)) for p in base.glob(pat)})
    except ValueError as e:
        return {"ok": False, "error": str(e), "paths": []}

    max_r = settings.GLOB_MAX_RESULTS
    truncated = len(matches) > max_r
    paths = matches[:max_r]
    return {"ok": True, "root": rel_root or ".", "pattern": pat, "paths": paths, "truncated": truncated}


def grep_tool(
    pattern: str,
    path: str = "",
    *,
    file_glob: str = "*",
    case_insensitive: bool = True,
) -> dict:
    """
    在工作区内搜索文本（Python re，不启 shell），对齐「grep 工具」需求、避免 bash 注入。
    path 为相对工作区的文件或目录；file_glob 仅对目录递归时过滤文件名（如 *.py）。
    """
    raw_pat = (pattern or "").strip()
    if not raw_pat or len(raw_pat) > 500:
        return {"ok": False, "error": "pattern 为空或过长", "matches": []}

    flags = re.MULTILINE
    if case_insensitive:
        flags |= re.IGNORECASE
    try:
        rx = re.compile(raw_pat, flags)
    except re.error as e:
        return {"ok": False, "error": f"正则无效: {e}", "matches": []}

    root_ws = _workspace_root()
    rel = (path or "").strip().replace("\\", "/").strip("/")
    try:
        target = (root_ws / rel).resolve() if rel else root_ws
        target.relative_to(root_ws)
    except ValueError:
        return {"ok": False, "error": "path 必须在工作区内", "matches": []}

    if not target.exists():
        return {"ok": False, "error": "路径不存在", "matches": []}

    skip_dirs = _grep_skip_dir_names()
    max_files = settings.GREP_MAX_FILES
    max_matches = settings.GREP_MAX_MATCHES
    matches: list[dict[str, str | int]] = []
    files_seen = 0

    def _grep_file(fp: Path) -> None:
        nonlocal files_seen
        if files_seen >= max_files or len(matches) >= max_matches:
            return
        if not fp.is_file():
            return
        if not fnmatch.fnmatch(fp.name, file_glob):
            return
        files_seen += 1
        try:
            raw = fp.read_bytes()[: settings.READ_FILE_MAX_BYTES]
        except OSError:
            return
        if b"\x00" in raw[:4096]:
            return
        try:
            text = raw.decode("utf-8", errors="strict")
        except UnicodeDecodeError:
            return
        for i, line in enumerate(text.splitlines(), start=1):
            if len(matches) >= max_matches:
                return
            if rx.search(line):
                rel_path = _rel_to_root(fp)
                snippet = line if len(line) <= 400 else line[:400] + "…"
                matches.append({"path": rel_path, "line": i, "text": snippet})

    if target.is_file():
        _grep_file(target)
    else:
        for dirpath, dirnames, filenames in os.walk(target, topdown=True):
            dp = Path(dirpath)
            try:
                dp.relative_to(root_ws)
            except ValueError:
                continue
            dirnames[:] = [d for d in dirnames if d not in skip_dirs and not d.startswith(".")]
            for name in sorted(filenames):
                if len(matches) >= max_matches or files_seen >= max_files:
                    break
                fp = dp / name
                _grep_file(fp)

    return {
        "ok": True,
        "pattern": raw_pat,
        "path": rel or ".",
        "file_glob": file_glob,
        "match_count": len(matches),
        "truncated": len(matches) >= max_matches or files_seen >= max_files,
        "matches": matches,
    }


def list_files_tool(path: str = "") -> dict:
    """actone `list_files`：列出工作区内某目录下的文件与子目录。"""
    root = _workspace_root()
    rel = (path or "").strip().replace("\\", "/").strip("/")
    try:
        target = (root / rel).resolve() if rel else root
        target.relative_to(root)
    except ValueError:
        return {"ok": False, "error": "路径必须在项目工作区内", "path": rel, "listing": ""}

    if not target.exists():
        return {"ok": False, "error": f"目录不存在: {rel or '.'}", "path": rel or ".", "listing": ""}
    if not target.is_dir():
        return {"ok": False, "error": "不是目录", "path": rel or ".", "listing": ""}

    items: list[str] = []
    dir_count = 0
    file_count = 0
    for p in sorted(target.iterdir()):
        if p.name.startswith("."):
            continue
        if p.is_dir():
            dir_count += 1
            try:
                nchild = len([c for c in p.iterdir() if not c.name.startswith(".")])
            except OSError:
                nchild = 0
            items.append(f"  📁 {p.name}/ ({nchild} items)")
        elif p.is_file():
            file_count += 1
            try:
                sz = p.stat().st_size
            except OSError:
                sz = 0
            size_str = f"{sz}B" if sz < 1024 else f"{sz / 1024:.1f}KB"
            items.append(f"  📄 {p.name} ({size_str})")

    header = f"📂 {rel or 'root'}: {dir_count} folder(s), {file_count} file(s)\n"
    listing = header + ("\n".join(items) if items else "(empty)")
    return {"ok": True, "path": rel or ".", "listing": listing}


def _protected_delete_set() -> set[str]:
    out: set[str] = set()
    for part in settings.WORKSPACE_DELETE_PROTECTED.split(","):
        p = part.strip().replace("\\", "/").lstrip("/")
        if p:
            out.add(p)
    return out


def write_file_tool(path: str, content: str) -> dict:
    """actone `write_file`：在工作区内写入或覆盖文本文件。"""
    if not settings.WORKSPACE_WRITE_ENABLED:
        return {"ok": False, "error": "WORKSPACE_WRITE_ENABLED=0，禁止写文件", "path": path or ""}

    raw = (content or "")
    max_b = settings.WORKSPACE_MAX_WRITE_BYTES
    if len(raw.encode("utf-8")) > max_b:
        return {"ok": False, "error": f"内容超过 WORKSPACE_MAX_WRITE_BYTES={max_b}", "path": path or ""}

    if not path or not str(path).strip():
        return {"ok": False, "error": "path 为空", "path": ""}

    try:
        fp = resolve_workspace_path(path)
    except ValueError as e:
        return {"ok": False, "error": str(e), "path": path}

    if _blocked_device_path(fp):
        return {"ok": False, "error": "拒绝写入该路径", "path": path}

    try:
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(raw, encoding="utf-8", newline="\n")
    except OSError as e:
        return {"ok": False, "error": str(e), "path": _rel_to_root(fp)}

    return {"ok": True, "path": _rel_to_root(fp), "bytes_written": len(raw.encode("utf-8"))}


def delete_file_tool(path: str) -> dict:
    """actone `delete_file`：删除工作区内单个文件（非目录）。"""
    if not settings.WORKSPACE_WRITE_ENABLED:
        return {"ok": False, "error": "WORKSPACE_WRITE_ENABLED=0，禁止删除", "path": path or ""}

    if not path or not str(path).strip():
        return {"ok": False, "error": "path 为空", "path": ""}

    try:
        fp = resolve_workspace_path(path)
    except ValueError as e:
        return {"ok": False, "error": str(e), "path": path}

    rel = _rel_to_root(fp)
    if rel in _protected_delete_set():
        return {"ok": False, "error": f"受保护路径，禁止删除: {rel}", "path": rel}

    if not fp.exists():
        return {"ok": False, "error": "文件不存在", "path": rel}
    if fp.is_dir():
        return {"ok": False, "error": "请只删除文件，不要删除目录", "path": rel}

    try:
        fp.unlink()
    except OSError as e:
        return {"ok": False, "error": str(e), "path": rel}

    return {"ok": True, "path": rel}


def read_document_tool(path: str) -> dict:
    """
    actone `read_document`：从 PDF / DOCX / XLSX 抽取文本（需可选依赖，未安装时返回明确提示）。
    .md/.txt/.json 等仍建议用 read_file。
    """
    if not path or not str(path).strip():
        return {"ok": False, "error": "path 为空", "path": "", "content": ""}

    try:
        fp = resolve_workspace_path(path)
    except ValueError as e:
        return {"ok": False, "error": str(e), "path": path, "content": ""}

    if not fp.is_file():
        return {"ok": False, "error": "不是文件", "path": _rel_to_root(fp), "content": ""}

    ext = fp.suffix.lower()
    max_c = settings.READ_DOCUMENT_MAX_CHARS

    try:
        if ext == ".pdf":
            try:
                import pdfplumber
            except ImportError:
                return {
                    "ok": False,
                    "error": "未安装 pdfplumber，无法读 PDF。请: pip install pdfplumber",
                    "path": _rel_to_root(fp),
                    "content": "",
                }
            parts: list[str] = []
            with pdfplumber.open(str(fp)) as pdf:
                for i, page in enumerate(pdf.pages[:50]):
                    t = page.extract_text() or ""
                    if t.strip():
                        parts.append(f"--- Page {i + 1} ---\n{t}")
            content = "\n\n".join(parts) if parts else "(empty or unreadable PDF)"

        elif ext == ".docx":
            try:
                from docx import Document
            except ImportError:
                return {
                    "ok": False,
                    "error": "未安装 python-docx。请: pip install python-docx",
                    "path": _rel_to_root(fp),
                    "content": "",
                }
            doc = Document(str(fp))
            lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            content = "\n".join(lines) if lines else "(empty)"

        elif ext == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                return {
                    "ok": False,
                    "error": "未安装 openpyxl。请: pip install openpyxl",
                    "path": _rel_to_root(fp),
                    "content": "",
                }
            wb = load_workbook(str(fp), read_only=True, data_only=True)
            chunks: list[str] = []
            for ws_name in wb.sheetnames[:10]:
                sheet = wb[ws_name]
                rows: list[str] = []
                for row in sheet.iter_rows(max_row=200, values_only=True):
                    line = "\t".join(str(c) if c is not None else "" for c in row)
                    if line.strip():
                        rows.append(line)
                if rows:
                    chunks.append(f"--- {ws_name} ---\n" + "\n".join(rows))
            content = "\n\n".join(chunks) if chunks else "(empty)"

        elif ext in (".md", ".txt", ".json", ".yaml", ".yml", ".csv"):
            content = fp.read_text(encoding="utf-8", errors="replace")

        else:
            return {
                "ok": False,
                "error": f"不支持的扩展名: {ext}，请用 read_file（文本）或安装对应依赖",
                "path": _rel_to_root(fp),
                "content": "",
            }

    except Exception as e:
        return {"ok": False, "error": str(e)[:500], "path": _rel_to_root(fp), "content": ""}

    full_len = len(content)
    truncated = full_len > max_c
    if truncated:
        content = content[:max_c] + f"\n\n...[truncated, total {full_len} chars]"

    return {"ok": True, "path": _rel_to_root(fp), "content": content, "truncated": truncated}


def read_skill_tool(skill_id: str) -> dict:
    """Load SKILL.md by relative path or basename directory under skill roots."""
    sid = (skill_id or "").strip()
    if not sid:
        return {"ok": False, "error": "skill_id 为空", "id": "", "content": ""}

    root = _workspace_root()
    # Direct path
    try:
        fp = resolve_workspace_path(sid)
        if fp.is_file() and fp.name == "SKILL.md":
            try:
                body = fp.read_text(encoding="utf-8", errors="replace")
            except OSError as e:
                return {"ok": False, "error": str(e), "id": sid, "content": ""}
            rel = _rel_to_root(fp)
            return {"ok": True, "id": rel, "content": body}

        if fp.is_dir():
            cand = fp / "SKILL.md"
            if cand.is_file():
                body = cand.read_text(encoding="utf-8", errors="replace")
                rel = _rel_to_root(cand)
                return {"ok": True, "id": rel, "content": body}
    except ValueError:
        pass

    # Search by directory name
    for base in _skills_dirs():
        for p in base.rglob("SKILL.md"):
            if p.parent.name == sid or str(p.relative_to(root)) == sid:
                try:
                    body = p.read_text(encoding="utf-8", errors="replace")
                    rel = str(p.relative_to(root))
                    return {"ok": True, "id": rel, "content": body}
                except OSError as e:
                    return {"ok": False, "error": str(e), "id": sid, "content": ""}

    return {"ok": False, "error": f"未找到 skill: {sid}", "id": sid, "content": ""}
